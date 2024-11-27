import os
import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def makeSalariesExcelFile(salaries_data: dict) -> str:
    '''Makes Excel file with employees salaries.
    
    :param salaries_data - dict with city, year, month and salaries_blocks
    '''

    city = salaries_data['city']
    year = salaries_data['year']
    month = salaries_data['month']
    salaries_blocks = salaries_data['salaries_blocks']

    filename = f"salaries-{city}-{year}-{month}.xlsx"
    file_path = f'salaries/salaries_xlsx_files/{filename}'

    if os.path.exists(file_path):
        file_ctime = datetime.datetime.fromtimestamp(os.path.getctime(file_path))
        if  (datetime.datetime.now() - file_ctime).seconds < 60**2:
            return filename

    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Отчёт"

    # Table styles
    header_font = Font(bold=True)
    block_font = Font(bold=True, color="FFFFFF") # White font
    block_fill = PatternFill(start_color="4F4F4F", end_color="4F4F4F", fill_type="solid")  # Dark-gray background
    metrics_fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")  # Gray background
    data_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")  # Light-gray background
    center_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000")
    )

    # Make Excel data
    row_num = 1
    for block in salaries_blocks:
        # 1. Block title
        block_title = block['block_data']['title']
        cell = ws.cell(row=row_num, column=1, value=block_title)
        cell.font = block_font
        cell.fill = block_fill
        cell.border = thin_border
        row_num += 1

        # 2. Table heading - metrics titles
        metrics_titles = (
            ["Сотрудник"] + [metric['title'] for metric in block['metrics_data'] if metric['id'] != 'employee']
        )
        for col_num, title in enumerate(metrics_titles, start=1):
            cell = ws.cell(row=row_num, column=col_num, value=title)
            cell.font = header_font
            cell.fill = metrics_fill
            cell.alignment = center_alignment
            cell.border = thin_border
        row_num += 1

        # 3. Table rows - employees metrics values
        for employee in block['employees']:
            # Employee fullname
            row_start_cell = ws.cell(row=row_num, column=1, value=employee['fullname'])
            row_start_cell.fill = data_fill
            row_start_cell.border = thin_border

            # Employee metrics values
            metrics = {m['id']: m['value'] for m in employee['metrics']['main']}
            for col_num, metric_id \
                in enumerate([m['id'] for m in block['metrics_data'] if m['id'] != 'employee'], start=2):
                    cell = ws.cell(row=row_num, column=col_num, value=round(metrics.get(metric_id, ""), 2))
                    cell.fill = data_fill  # Light-gray background
                    cell.border = thin_border
            row_num += 1

        # 4. Empty row between blocks
        row_num += 1
        

    # Columns width configuration
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 2

    wb.save(file_path)

    return filename